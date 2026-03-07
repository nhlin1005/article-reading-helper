# lookup_single_word.py
# -*- coding: utf-8 -*-
"""
Standalone single-word lookup.
Uses the exact same XLSX loading logic as build_vocab_combined.py so it finds
the same columns (word / meaning / example) regardless of what they're called.

Also falls back to the SQLite cache that build_vocab_combined already populates,
then to Merriam-Webster + Cambridge scraping as a last resort.
"""

import re
import os
import sqlite3
import time
import random
from pathlib import Path
from typing import Optional, Dict, Tuple

# ── same column-candidate lists as build_vocab_combined ─────────────────────
WORD_COL_CANDIDATES    = ["word", "term", "vocab", "vocabulary", "lemma", "headword"]
MEANING_COL_CANDIDATES = ["definitionen", "meaning", "definition", "gloss", "def"]
EXAMPLE_COL_CANDIDATES = ["example", "examples", "sentence", "sentences", "eg"]


def _canon(s: str) -> str:
    """Normalise a header name the same way build_vocab_combined does."""
    return re.sub(r"[^a-z0-9]+", "", (s or "").strip().lower())


# ── XLSX loading ─────────────────────────────────────────────────────────────
def _load_xlsx_lookup(xlsx_path: str, sheet_name: Optional[str] = "All_Words") -> Dict[str, Tuple[str, str]]:
    """
    Load the XLSX dictionary into {word_lower: (meaning, example)}.
    Identical logic to build_vocab_combined.load_xlsx_lookup().
    """
    if not xlsx_path or not Path(xlsx_path).exists():
        return {}

    try:
        import openpyxl
    except ImportError:
        return {}

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:
        return {}

    ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active

    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(c.value).strip() if c.value is not None else "" for c in header_cells]
    canon_to_idx = {_canon(h): i for i, h in enumerate(headers)}

    def find_col(cands):
        for c in cands:
            k = _canon(c)
            if k in canon_to_idx:
                return canon_to_idx[k]
        return None

    word_i    = find_col(WORD_COL_CANDIDATES)
    meaning_i = find_col(MEANING_COL_CANDIDATES)
    example_i = find_col(EXAMPLE_COL_CANDIDATES)

    if word_i is None:
        wb.close()
        return {}

    lookup: Dict[str, Tuple[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        w = row[word_i] if (row and word_i < len(row)) else None
        if w is None:
            continue
        w = str(w).strip()
        if not w:
            continue

        meaning = ""
        example = ""
        if meaning_i is not None and meaning_i < len(row) and row[meaning_i] is not None:
            meaning = str(row[meaning_i]).strip()
        if example_i is not None and example_i < len(row) and row[example_i] is not None:
            example = str(row[example_i]).strip()

        if meaning or example:
            lookup[w.lower()] = (meaning, example)

    wb.close()
    return lookup


# Cache the loaded XLSX in memory so repeated lookups are fast
_xlsx_cache: Optional[Dict[str, Tuple[str, str]]] = None
_xlsx_path_loaded: str = ""


def _get_xlsx_lookup(xlsx_path: str) -> Dict[str, Tuple[str, str]]:
    global _xlsx_cache, _xlsx_path_loaded
    if _xlsx_cache is None or _xlsx_path_loaded != xlsx_path:
        _xlsx_cache = _load_xlsx_lookup(xlsx_path, sheet_name="All_Words")
        _xlsx_path_loaded = xlsx_path
    return _xlsx_cache


# ── SQLite cache (same DB that build_vocab_combined uses) ────────────────────
def _cache_get(db_path: str, word: str) -> Tuple[str, str]:
    if not db_path or not Path(db_path).exists():
        return "", ""
    try:
        conn = sqlite3.connect(db_path, timeout=5)
        cur  = conn.execute("SELECT meaning, example FROM vocab_cache WHERE word=?", (word.lower(),))
        row  = cur.fetchone()
        conn.close()
        return (row[0] or "", row[1] or "") if row else ("", "")
    except Exception:
        return "", ""


# ── Simple stem variants (same as before) ────────────────────────────────────
def _stems(word: str):
    yield word
    if word.endswith("ing") and len(word) > 5:
        base = word[:-3]
        yield base
        yield base + "e"
    if word.endswith("ed") and len(word) > 4:
        base = word[:-2]
        yield base
        yield base + "e"
    if word.endswith("es") and len(word) > 4:
        yield word[:-2]
        yield word[:-1]
    elif word.endswith("s") and len(word) > 3:
        yield word[:-1]
    if word.endswith("ly") and len(word) > 4:
        yield word[:-2]
    if word.endswith("tion") and len(word) > 6:
        yield word[:-4]
        yield word[:-3] + "e"


# ── Web scraping fallbacks (Merriam-Webster + Cambridge) ─────────────────────
def _webster_meaning(word: str) -> str:
    try:
        import requests
        from lxml import html as lhtml
        from urllib.parse import quote
        url = f"https://www.merriam-webster.com/dictionary/{quote(word)}"
        r = requests.get(url, timeout=10, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        })
        if r.status_code != 200:
            return ""
        tree = lhtml.fromstring(r.text)
        nodes = tree.xpath(
            "//div[contains(@id,'dictionary-entry')]/div[@class='vg']"
            "//span[@class='dtText' or @class='unText']"
        )
        for n in nodes:
            txt = re.sub(r"<.*?>", "", lhtml.tostring(n, encoding="unicode"))
            txt = re.sub(r"^\s*:\s*", "", txt).strip()
            if txt:
                return txt
    except Exception:
        pass
    return ""


def _cambridge_example(word: str) -> str:
    try:
        import requests
        from lxml import html as lhtml
        from urllib.parse import quote
        urls = [
            f"https://dictionary.cambridge.org/dictionary/english/{quote(word)}",
            f"https://dictionary.cambridge.org/us/dictionary/english/{quote(word)}",
        ]
        xpaths = [
            "//div[@class='def-body']//span[@class='eg']",
            "//div[@class='examp dexamp']//span",
            "//span[contains(@class,'eg') and not(contains(@class,'dsense'))]",
        ]
        for url in urls:
            r = requests.get(url, timeout=10, headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            })
            if r.status_code != 200:
                continue
            tree = lhtml.fromstring(r.text)
            for xp in xpaths:
                nodes = tree.xpath(xp)
                for n in nodes:
                    text = re.sub(r"\s+", " ", "".join(n.xpath(".//text()")).strip())
                    if len(text) > 20 and len(text.split()) >= 4:
                        return text
    except Exception:
        pass
    return ""


# ── Public API ────────────────────────────────────────────────────────────────
def lookup_word_in_xlsx(word: str, xlsx_path: str, cache_db: str = "") -> dict:
    """
    Look up a word. Priority:
      1. XLSX dictionary (exact match, then stem variants)
      2. SQLite cache (if cache_db provided)
      3. Web scrape: Merriam-Webster meaning + Cambridge example

    Returns: { word, meaning, example, ipa }
    """
    word = (word or "").strip()
    if not word:
        return {"word": word, "meaning": "", "example": "", "ipa": ""}

    lookup = _get_xlsx_lookup(xlsx_path)
    word_lower = word.lower()

    # 1. XLSX — exact then stem
    for stem in _stems(word_lower):
        if stem in lookup:
            meaning, example = lookup[stem]
            return {"word": word, "meaning": meaning, "example": example, "ipa": ""}

    # 2. SQLite cache
    if cache_db:
        meaning, example = _cache_get(cache_db, word_lower)
        if meaning or example:
            return {"word": word, "meaning": meaning, "example": example, "ipa": ""}

    # 3. Web scrape
    meaning = _webster_meaning(word)
    example = _cambridge_example(word)

    return {"word": word, "meaning": meaning, "example": example, "ipa": ""}