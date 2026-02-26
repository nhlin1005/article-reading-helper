# -*- coding: utf-8 -*-
"""
pipeline_from_pdf.py

新流程（AI 模式）：
1) PDF -> 文章 txt
2) 用 XLSX 词库把“文章里出现且在词库里”的词全部选出来（按出现顺序去重）
3) 对剩下的 OOV（不在词库里）部分再做 AI 选词补充
4) 最终 select = in_vocab + ai_selected_oov
5) 调用 build_vocab 生成 words.txt + csv + json（build_vocab 内部应“优先 xlsx，缺失再 Merriam-Webster / Cambridge”）

list 模式：
- 仍然使用你提供的 select.txt，但 build_vocab 调用仍会传 lookup_xlsx/xlsx_sheet（若 build_vocab 不支持会自动降级）

依赖：
- openpyxl：用于读取 xlsx 词库（没有也能跑，但 AI 模式就无法“先词库命中”）
"""

import argparse
import json
import math
import re
from pathlib import Path
from collections import Counter

from extract_pdf_text import extract_text_from_pdf
from build_vocab_combined import build_vocab
from csv_to_json import csv_to_json
from ai_select_wordlist import ai_select_words_for_article
from config import DEFAULT_TOP_N

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_XLSX = BASE_DIR / "vocab" / "words_with_examples.xlsx"
# ----------------------------
# Folder naming
# ----------------------------
def _make_reading_folder(pdf_path: Path):
    pdf_stem = pdf_path.stem
    safe_stem = "".join(c if (c.isalnum() or c in "-_") else "_" for c in pdf_stem)
    reading_dir = Path(f"reading_{safe_stem}")
    reading_dir.mkdir(parents=True, exist_ok=True)
    article_txt_path = reading_dir / f"{safe_stem}.txt"
    return reading_dir, safe_stem, article_txt_path


# ----------------------------
# Difficulty score JSON (optional)
# ----------------------------
def _normalize_for_freq(s: str) -> str:
    return (s or "").strip().lower()


def _make_score_json(article_text: str, words: list[str]) -> dict:
    toks = [_normalize_for_freq(w) for w in article_text.split()]
    freq = Counter(toks)

    max_f = 1
    for w in words:
        max_f = max(max_f, freq.get(_normalize_for_freq(w), 1))

    scores = {}
    for w in words:
        ww = _normalize_for_freq(w)
        f = max(1, freq.get(ww, 1))
        rare_score = 1.0 - (math.log(f + 1.0) / math.log(max_f + 1.0))
        len_score = min(1.0, len(ww) / 12.0)
        s = 0.65 * rare_score + 0.35 * len_score
        scores[w] = round(float(s), 3)

    return scores


# ----------------------------
# XLSX vocab pre-selection
# ----------------------------
_WORD_RE = re.compile(r"[A-Za-z][A-Za-z\-']*")


def _canon(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (s or "").strip().lower())


def _unique_preserve_order(items: list[str]) -> list[str]:
    seen = set()
    out = []
    for x in items:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


def _tokenize_article_words(text: str) -> list[str]:
    return [m.group(0).lower() for m in _WORD_RE.finditer(text or "")]


def _load_vocab_set_from_xlsx(xlsx_path: str, sheet: str = "All_Words") -> set[str]:
    """
    Load a set of words from XLSX.
    Looks for a header like Word/word/term/vocab... otherwise uses the first column.
    """
    try:
        import openpyxl
    except Exception:
        print("WARNING: openpyxl not installed; cannot pre-select by XLSX vocab. "
              "Run: python -m pip install openpyxl")
        return set()

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"WARNING: cannot open xlsx {xlsx_path}: {e}")
        return set()

    ws = wb[sheet] if (sheet and sheet in wb.sheetnames) else wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(v).strip() if v is not None else "" for v in header_row]
    hmap = {_canon(h): i for i, h in enumerate(headers)}

    word_i = None
    for k in ["word", "term", "vocab", "vocabulary", "headword", "lemma"]:
        ck = _canon(k)
        if ck in hmap:
            word_i = hmap[ck]
            break
    if word_i is None:
        word_i = 0  # fallback: first column

    vocab = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or word_i >= len(row):
            continue
        v = row[word_i]
        if v is None:
            continue
        w = str(v).strip().lower()
        if w:
            vocab.add(w)

    return vocab


def _split_by_vocab(article_tokens: list[str], vocab_set: set[str]) -> tuple[list[str], list[str]]:
    ordered = _unique_preserve_order(article_tokens)
    in_vocab = [w for w in ordered if w in vocab_set]
    oov = [w for w in ordered if w not in vocab_set]
    return in_vocab, oov


# ----------------------------
# build_vocab call compatibility
# ----------------------------
def _call_build_vocab(article_txt: Path,
                      select_path: Path,
                      out_words: Path,
                      out_csv: Path,
                      score_json_path: str,
                      lookup_xlsx: str,
                      xlsx_sheet: str):
    """
    Try newer signatures first, then fall back.
      - build_vocab(a, s, ow, oc, score_json, lookup_xlsx, xlsx_sheet)
      - build_vocab(a, s, ow, oc, lookup_xlsx, xlsx_sheet)
      - build_vocab(a, s, ow, oc, score_json)
      - build_vocab(a, s, ow, oc)
    """
    a = str(article_txt)
    s = str(select_path)
    ow = str(out_words)
    oc = str(out_csv)

    try:
        return build_vocab(a, s, ow, oc, score_json_path, lookup_xlsx, xlsx_sheet)
    except TypeError:
        pass

    try:
        return build_vocab(a, s, ow, oc, lookup_xlsx, xlsx_sheet)
    except TypeError:
        pass

    try:
        return build_vocab(a, s, ow, oc, score_json_path)
    except TypeError:
        pass

    return build_vocab(a, s, ow, oc)


# ----------------------------
# Modes
# ----------------------------
def run_ai_mode(pdf_path: Path,
                reading_dir: Path,
                safe_stem: str,
                ai_top_n: float = DEFAULT_TOP_N,
                lookup_xlsx: str = "words_with_examples.xlsx",
                xlsx_sheet: str = "All_Words"):

    # Step 1: PDF -> TXT
    article_txt = reading_dir / f"{safe_stem}.txt"
    print(f"\n[Step 1] Extract PDF text: {pdf_path} -> {article_txt}")
    ok = extract_text_from_pdf(str(pdf_path), str(article_txt))
    if not ok:
        raise SystemExit("❌ PDF text extraction failed.")

    text = article_txt.read_text(encoding="utf-8", errors="ignore")

    # Step 2: vocab pre-select + AI补充
    select_path = reading_dir / f"{safe_stem}.ai.select.txt"

    vocab_set = _load_vocab_set_from_xlsx(lookup_xlsx, sheet=xlsx_sheet)
    if vocab_set:
        print(f"\n[Step 2] Loaded vocab from XLSX: {lookup_xlsx} (sheet={xlsx_sheet}) "
              f"-> {len(vocab_set)} words")
    else:
        print(f"\n[Step 2] XLSX vocab unavailable or empty; AI will select from full article.")

    article_tokens = _tokenize_article_words(text)

    if vocab_set:
        in_vocab_words, oov_words = _split_by_vocab(article_tokens, vocab_set)
        print(f"  - In-vocab words found in article: {len(in_vocab_words)}")
        print(f"  - OOV words (not in vocab): {len(oov_words)}")

        # AI only sees OOV words (as simplified pseudo-article)
        oov_text_for_ai = " ".join(oov_words)
        ai_raw = ai_select_words_for_article(oov_text_for_ai, top_n=ai_top_n) if oov_words else []
        ai_set = set([w.lower() for w in ai_raw])

        # Keep AI picks in original article order (OOV order)
        ai_words_ordered = [w for w in oov_words if w in ai_set]

        words = in_vocab_words + ai_words_ordered
        print(f"  ✓ Final selection: vocab {len(in_vocab_words)} + AI(OOV) {len(ai_words_ordered)} = {len(words)}")
    else:
        # fallback: original AI behavior on full text
        ai_raw = ai_select_words_for_article(text, top_n=ai_top_n)
        # order them by first appearance in article
        ordered = _unique_preserve_order(article_tokens)
        ai_set = set([w.lower() for w in ai_raw])
        words = [w for w in ordered if w in ai_set]
        print(f"  ✓ Final selection (AI only): {len(words)}")

    select_path.write_text("\n".join(words) + "\n", encoding="utf-8")
    print(f"  ✓ Saved select list: {select_path}")

    # Step 2.5: difficulty score JSON
    score_json_path = reading_dir / f"{safe_stem}.selected_words.json"
    scores = _make_score_json(text, words)
    score_json_path.write_text(json.dumps(scores, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  ✓ Saved difficulty scores: {score_json_path}")

    # Step 3: build vocab outputs
    out_words = reading_dir / f"{safe_stem}.ai.words.txt"
    out_csv = reading_dir / f"{safe_stem}.ai.csv"

    print(f"\n[Step 3] Build vocab CSV (XLSX-first; fallback to dictionaries)...")
    _call_build_vocab(
        article_txt=article_txt,
        select_path=select_path,
        out_words=out_words,
        out_csv=out_csv,
        score_json_path=str(score_json_path),
        lookup_xlsx=lookup_xlsx,
        xlsx_sheet=xlsx_sheet,
    )

    print(f"  ✓ Out words: {out_words}")
    print(f"  ✓ Out CSV: {out_csv}")

    # Step 4: CSV -> JSON
    out_json = reading_dir / f"{safe_stem}.ai.json"
    print(f"\n[Step 4] CSV -> JSON: {out_csv} -> {out_json}")
    csv_to_json(str(out_csv), str(out_json))
    print(f"  ✓ Out JSON: {out_json}")

    print("\n✅ AI mode done. Output folder:", reading_dir)


def run_list_mode(pdf_path: Path,
                  reading_dir: Path,
                  safe_stem: str,
                  select_path: Path,
                  lookup_xlsx: str = "words_with_examples.xlsx",
                  xlsx_sheet: str = "All_Words"):

    if not select_path.exists():
        raise SystemExit(f"❌ select file not found: {select_path}")

    # Step 1: PDF -> TXT
    article_txt = reading_dir / f"{safe_stem}.txt"
    print(f"\n[Step 1] Extract PDF text: {pdf_path} -> {article_txt}")
    ok = extract_text_from_pdf(str(pdf_path), str(article_txt))
    if not ok:
        raise SystemExit("❌ PDF text extraction failed.")

    # Step 2: build vocab
    out_words = reading_dir / f"{safe_stem}.list.words.txt"
    out_csv = reading_dir / f"{safe_stem}.list.csv"

    print(f"\n[Step 2] Build vocab CSV using your select list (XLSX-first; fallback to dictionaries)...")
    _call_build_vocab(
        article_txt=article_txt,
        select_path=select_path,
        out_words=out_words,
        out_csv=out_csv,
        score_json_path="",
        lookup_xlsx=lookup_xlsx,
        xlsx_sheet=xlsx_sheet,
    )

    print(f"  ✓ Out words: {out_words}")
    print(f"  ✓ Out CSV: {out_csv}")

    # Step 3: CSV -> JSON
    out_json = reading_dir / f"{safe_stem}.list.json"
    print(f"\n[Step 3] CSV -> JSON: {out_csv} -> {out_json}")
    csv_to_json(str(out_csv), str(out_json))
    print(f"  ✓ Out JSON: {out_json}")

    print("\n✅ List mode done. Output folder:", reading_dir)


# ----------------------------
# CLI
# ----------------------------
def main():
    ap = argparse.ArgumentParser(description="PDF -> select -> vocab CSV/JSON (XLSX vocab preselect + AI补充).")
    ap.add_argument("--mode", choices=["ai", "list"], required=True,
                    help="ai = vocab命中 + AI补充；list = 用你提供的select.txt")
    ap.add_argument("--pdf", required=True, help="Input PDF path")

    ap.add_argument("--select", default=None,
                    help="When --mode list, path to select .txt")

    ap.add_argument("--ai_top_n", type=float, default=DEFAULT_TOP_N,
                    help="AI picks: >=1 means count; (0,1] means ratio (depends on your ai_select_words_for_article).")

    ap.add_argument("--lookup_xlsx", default=str(DEFAULT_XLSX),
                    help="XLSX vocab/lookup file (default: words_with_examples.xlsx)")
    ap.add_argument("--xlsx_sheet", default="All_Words",
                    help="XLSX sheet name (default: All_Words)")

    args = ap.parse_args()

    pdf_path = Path(args.pdf)
    if not pdf_path.exists():
        raise SystemExit(f"❌ PDF not found: {pdf_path}")

    reading_dir, safe_stem, article_txt_path = _make_reading_folder(pdf_path)
    print(f"📁 Output folder: {reading_dir}")
    print(f"📚 XLSX vocab: {args.lookup_xlsx} (sheet={args.xlsx_sheet})")

    if args.mode == "ai":
        run_ai_mode(
            pdf_path=pdf_path,
            reading_dir=reading_dir,
            safe_stem=safe_stem,
            ai_top_n=args.ai_top_n,
            lookup_xlsx=args.lookup_xlsx,
            xlsx_sheet=args.xlsx_sheet,
        )
    else:
        if not args.select:
            raise SystemExit("❌ --mode list requires --select <select.txt>")
        run_list_mode(
            pdf_path=pdf_path,
            reading_dir=reading_dir,
            safe_stem=safe_stem,
            select_path=Path(args.select),
            lookup_xlsx=args.lookup_xlsx,
            xlsx_sheet=args.xlsx_sheet,
        )


if __name__ == "__main__":
    main()
