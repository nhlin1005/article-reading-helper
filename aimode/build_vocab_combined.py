# -*- coding: utf-8 -*-
"""
Builds a vocabulary CSV from an article + a selection list of words.

Priority:
  - First try lookup from an XLSX file (meaning + example).
  - Then try local cache (sqlite) for missing fields.
  - If still missing meaning: MEANING from Merriam-Webster.
  - If still missing example: EXAMPLE from Cambridge Dictionary.
  - If Cambridge has no example: fallback to an article sentence containing the word.
  - If meaning still missing: heuristic/AI fallback.

Speed-ups:
  - SQLite cache to avoid repeated network calls across runs.
  - Limited concurrency for network fetching (ThreadPoolExecutor).
  - Backoff retry for 429/503/520-ish throttling.

Usage:
  python build_vocab_combined.py --article sat.txt --select 20241226.txt \
      --lookup_xlsx words_with_examples.xlsx --xlsx_sheet All_Words \
      --cache_db vocab_cache.sqlite --workers 8

Notes:
  - If you are being throttled, reduce --workers and increase --jitter_max.
"""

import argparse
import csv
import time
import string
import re
import random
import os
import sqlite3
from typing import Optional, List, Iterable, Dict, Tuple
from urllib.parse import quote
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from lxml import html

import functools
print = functools.partial(print, flush=True)

# ----------------------------
# Optional AI meaning/example fallback (LLM)
# ----------------------------
try:
    from llm_refiner import generate_meaning_example_with_context
except Exception:
    generate_meaning_example_with_context = None


# ----------------------------
# XLSX lookup (openpyxl)
# ----------------------------
try:
    import openpyxl
except Exception:
    openpyxl = None


def _canon(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (s or "").strip().lower())


WORD_COL_CANDIDATES = ["word", "term", "vocab", "vocabulary", "lemma", "headword"]
MEANING_COL_CANDIDATES = ["definitionen", "meaning", "definition", "gloss", "def"]
EXAMPLE_COL_CANDIDATES = ["example", "examples", "sentence", "sentences", "eg"]


def load_xlsx_lookup(xlsx_path: str, sheet_name: Optional[str]) -> Dict[str, Tuple[str, str]]:
    if not xlsx_path:
        return {}
    if openpyxl is None:
        print("WARNING: openpyxl not installed; XLSX lookup disabled. Run: python -m pip install openpyxl")
        return {}

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except FileNotFoundError:
        print(f"WARNING: lookup_xlsx not found: {xlsx_path} (XLSX lookup disabled)")
        return {}
    except Exception as e:
        print(f"WARNING: failed to open lookup_xlsx {xlsx_path}: {e}")
        return {}

    ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active

    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(c.value).strip() if c.value is not None else "" for c in header_cells]
    canon_to_idx = {_canon(h): i for i, h in enumerate(headers)}  # 0-based

    def find_col(cands: List[str]) -> Optional[int]:
        for c in cands:
            k = _canon(c)
            if k in canon_to_idx:
                return canon_to_idx[k]
        return None

    word_i = find_col(WORD_COL_CANDIDATES)
    meaning_i = find_col(MEANING_COL_CANDIDATES)
    example_i = find_col(EXAMPLE_COL_CANDIDATES)

    if word_i is None:
        print(f"WARNING: XLSX lookup: cannot find word column in sheet '{ws.title}'. Headers={headers}")
        return {}
    if meaning_i is None and example_i is None:
        print(f"WARNING: XLSX lookup: cannot find meaning/example columns in sheet '{ws.title}'. Headers={headers}")
        return {}

    lookup: Dict[str, Tuple[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        w = row[word_i] if (row and word_i < len(row)) else None
        if w is None:
            continue
        w = str(w).strip()
        if not w:
            continue
        key = w.lower()

        meaning = ""
        example = ""
        if meaning_i is not None and meaning_i < len(row) and row[meaning_i] is not None:
            meaning = str(row[meaning_i]).strip()
        if example_i is not None and example_i < len(row) and row[example_i] is not None:
            example = str(row[example_i]).strip()

        if meaning or example:
            lookup[key] = (meaning, example)

    print(f"✓ XLSX lookup loaded: {len(lookup)} entries from '{ws.title}' in {xlsx_path}")
    return lookup


# ----------------------------
# Live progress for front-end polling
# ----------------------------
PROGRESS = {"status": "idle", "current": 0, "total": 0, "word": ""}


def _set_progress(*, status: str = None, current: int = None, total: int = None, word: str = None) -> None:
    try:
        if status is not None:
            PROGRESS["status"] = status
        if current is not None:
            PROGRESS["current"] = int(current)
        if total is not None:
            PROGRESS["total"] = int(total)
        if word is not None:
            PROGRESS["word"] = word
    except Exception:
        pass


# ----------------------------
# Tokenization / normalization
# ----------------------------
PUNCTUATION = string.punctuation
WORD_RE = re.compile(r"[A-Za-z][A-Za-z\-']*")


def normalize_token(tok: str) -> str:
    tok = tok.strip().strip(PUNCTUATION)
    m = WORD_RE.match(tok)
    return m.group(0).lower() if m else tok.lower()


def tokenize_file(path: str) -> List[str]:
    out = []
    with open(path, "r", encoding="utf-8", errors="ignore") as fp:
        for line in fp:
            for raw in line.split():
                tok = normalize_token(raw)
                if tok:
                    out.append(tok)
    return out


# ----------------------------
# HTTP helpers
# ----------------------------
SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
})
TIMEOUT = 12


def fetch(url: str) -> Tuple[Optional[str], int]:
    try:
        r = SESSION.get(url, timeout=TIMEOUT)
        return r.text, r.status_code
    except requests.RequestException:
        return None, -1


def get_tree(url: str):
    text, code = fetch(url)
    if code != 200 or not text:
        return None, code
    try:
        return html.fromstring(text), code
    except Exception:
        return None, code


def normalize_space(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()


def _retry_sleep(attempt: int, base: float = 0.6, cap: float = 8.0) -> None:
    # exponential backoff + jitter
    t = min(cap, base * (2 ** attempt)) + random.uniform(0.0, 0.25)
    time.sleep(t)


# ----------------------------
# Merriam-Webster meaning
# ----------------------------
def _webster_meaning_core(word: str) -> Optional[str]:
    url = f"https://www.merriam-webster.com/dictionary/{quote(word)}"
    tree, code = get_tree(url)
    if tree is None:
        return None

    def_nodes = tree.xpath(
        "//div[contains(@id,'dictionary-entry')]/div[@class='vg']"
        "//span[@class='dtText' or @class='unText']"
    )
    for n in def_nodes:
        txt = html.tostring(n, encoding="unicode")
        txt = re.sub(r"<.*?>", "", txt)
        txt = re.sub(r"^\s*:\s*", "", txt).strip()
        if txt:
            return txt
    return None


def webster_meaning(word: str) -> Optional[str]:
    meaning = _webster_meaning_core(word)
    if meaning:
        return meaning

    if "'" not in word and "our" in word:
        alt = word.replace("our", "or")
        if alt != word:
            alt_meaning = _webster_meaning_core(alt)
            if alt_meaning:
                return alt_meaning
    return None


# ----------------------------
# Cambridge example
# ----------------------------
def cambridge_example(word: str) -> Optional[str]:
    urls = [
        f"https://dictionary.cambridge.org/dictionary/english/{quote(word)}",
        f"https://dictionary.cambridge.org/us/dictionary/english/{quote(word)}",
    ]

    xpaths = [
        "//div[@class='def-body']//span[@class='eg']",
        "//div[@class='examp dexamp']//span",
        "//span[contains(@class,'eg') and not(contains(@class,'dsense'))]",
        "//div[contains(@class,'example-box')]//span[@class='eg']",
        "//div[contains(@class,'examp')]//span[contains(@class,'eg')]",
    ]

    for url in urls:
        tree, code = get_tree(url)
        if tree is None:
            continue

        for xp in xpaths:
            try:
                nodes = tree.xpath(xp)
            except Exception:
                continue
            for n in nodes:
                text = normalize_space("".join(n.xpath(".//text()")))
                if (
                    len(text) > 20 and len(text.split()) >= 4
                    and not any(skip in text.lower() for skip in [
                        "more examples", "fewer examples", "smart vocabulary",
                        "thesaurus", "see also", "compare", "related to",
                    ])
                ):
                    return text
    return None


# ----------------------------
# Article example fallback
# ----------------------------
def find_sentence_as_example(article_text: str, word: str, max_len: int = 200) -> str:
    if not article_text:
        return ""
    pattern = re.compile(r"[^.?!]*\b" + re.escape(word) + r"\b[^.?!]*[.?!]", flags=re.IGNORECASE)
    m = pattern.search(article_text)
    if not m:
        return ""
    sent = normalize_space(m.group(0))
    if len(sent) > max_len:
        return sent[: max_len - 3].rstrip() + "..."
    return sent


def get_surrounding_context(article_text: str, word: str, window: int = 2, max_chars: int = 900) -> str:
    if not article_text or not word:
        return ""
    cleaned = normalize_space(article_text)
    if not cleaned:
        return ""
    sents = re.split(r"(?<=[.!?])\s+", cleaned)
    wpat = re.compile(r"\b" + re.escape(word) + r"\b", flags=re.IGNORECASE)

    hit = -1
    for i, s in enumerate(sents):
        if wpat.search(s):
            hit = i
            break

    if hit < 0:
        return cleaned[:max_chars]

    start_i = max(0, hit - window)
    end_i = min(len(sents), hit + window + 1)
    snippet = " ".join(sents[start_i:end_i]).strip()
    if len(snippet) > max_chars:
        snippet = snippet[: max_chars - 3].rstrip() + "..."
    return snippet


def heuristic_ai_meaning(word: str, context: str, example: str = "") -> str:
    w = (word or "").strip()
    if not w:
        return ""
    wl = w.lower()

    if "-" in wl:
        parts = [p for p in re.split(r"-+", wl) if p]
        if wl.endswith("-based") and len(parts) >= 2:
            base = " ".join(parts[:-1])
            return f"Based in or organized around {base}; used here as a descriptive modifier."
        if wl.endswith("-speaking") and len(parts) >= 2:
            lang = " ".join(parts[:-1])
            return f"Able to speak {lang}; used here to describe a person or group."
        if len(parts) == 2:
            return f"A compound adjective combining '{parts[0]}' and '{parts[1]}' in this context."
        return "A compound/hyphenated term used descriptively in this context."

    if example:
        return "A term used in the article; infer its meaning from the example sentence (may be specialized)."
    return "A term used in the article; it may be specialized or a proper noun not covered by standard dictionaries."


# ----------------------------
# Cache (SQLite)
# ----------------------------
def init_cache(db_path: str) -> sqlite3.Connection:
    os.makedirs(os.path.dirname(db_path), exist_ok=True) if os.path.dirname(db_path) else None
    conn = sqlite3.connect(db_path, timeout=30)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS vocab_cache (
            word TEXT PRIMARY KEY,
            meaning TEXT,
            example TEXT,
            updated_at INTEGER
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_vocab_cache_updated_at ON vocab_cache(updated_at)")
    conn.commit()
    return conn


def cache_get(conn: sqlite3.Connection, word: str) -> Tuple[str, str]:
    cur = conn.execute("SELECT meaning, example FROM vocab_cache WHERE word=?", (word.lower(),))
    row = cur.fetchone()
    if not row:
        return "", ""
    return (row[0] or ""), (row[1] or "")


def cache_put(conn: sqlite3.Connection, word: str, meaning: str, example: str) -> None:
    conn.execute(
        "INSERT INTO vocab_cache(word, meaning, example, updated_at) VALUES(?,?,?,?) "
        "ON CONFLICT(word) DO UPDATE SET meaning=excluded.meaning, example=excluded.example, updated_at=excluded.updated_at",
        (word.lower(), meaning or "", example or "", int(time.time()))
    )


# ----------------------------
# Helpers
# ----------------------------
def unique_preserve_order(items: Iterable[str]) -> List[str]:
    seen = set()
    out = []
    for x in items:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


def is_good_candidate(word: str) -> bool:
    if "'" in word:
        return False
    if word.count("-") >= 2:
        return False
    return True


def fetch_missing_fields_for_word(
    word: str,
    need_meaning: bool,
    need_example: bool,
    *,
    retries: int,
    jitter_min: float,
    jitter_max: float,
) -> Tuple[str, str]:
    """
    Network fetch only. Return (meaning_or_empty, example_or_empty).
    """
    meaning = ""
    example = ""

    # light jitter so concurrent workers don't synchronize
    if jitter_max > 0:
        time.sleep(random.uniform(jitter_min, jitter_max))

    for attempt in range(retries + 1):
        try:
            if need_meaning and not meaning:
                meaning = webster_meaning(word) or ""

            if need_example and not example:
                example = cambridge_example(word) or ""

            # success or no more needed
            return meaning, example
        except Exception:
            # in case parsing throws
            if attempt < retries:
                _retry_sleep(attempt)
            else:
                return meaning, example

    return meaning, example


# ----------------------------
# Main build function
# ----------------------------
def build_vocab(
    article_path: str,
    select_path: str,
    out_words: str,
    out_csv: str,
    lookup_xlsx: str,
    xlsx_sheet: Optional[str],
    *,
    cache_db: str = "vocab_cache.sqlite",
    workers: int = 8,
    retries: int = 2,
    jitter_min: float = 0.05,
    jitter_max: float = 0.25,
) -> None:

    # read full article for fallback examples / context
    try:
        with open(article_path, "r", encoding="utf-8", errors="ignore") as fp:
            article_text = fp.read()
    except Exception:
        article_text = ""

    # load xlsx lookup once
    xlsx_lookup = load_xlsx_lookup(lookup_xlsx, xlsx_sheet)

    # cache init
    cache_conn = init_cache(cache_db)

    article_tokens = tokenize_file(article_path)
    selection_tokens = tokenize_file(select_path)

    filtered_selection = [w for w in selection_tokens if is_good_candidate(w)]
    selected = set(filtered_selection)

    ordered = [w for w in article_tokens if w in selected]
    ordered = unique_preserve_order(ordered)

    _set_progress(status="running", current=0, total=len(ordered), word="")
    print(f"Processing {len(ordered)} words...\n")

    # write words list
    with open(out_words, "w", encoding="utf-8") as f:
        for w in ordered:
            f.write(f"{w}\n")

    # Phase 1: resolve from XLSX + cache, mark which need network
    rows: List[Tuple[str, str, str]] = []
    # store intermediate results
    meaning_map: Dict[str, str] = {}
    example_map: Dict[str, str] = {}
    needs_net: List[Tuple[str, bool, bool]] = []

    for w in ordered:
        meaning = ""
        example = ""

        # XLSX first
        x_meaning, x_example = xlsx_lookup.get(w.lower(), ("", ""))
        if x_meaning:
            meaning = x_meaning
        if x_example:
            example = x_example

        # cache second (only fill missing)
        if cache_conn:
            c_meaning, c_example = cache_get(cache_conn, w)
            if not meaning and c_meaning:
                meaning = c_meaning
            if not example and c_example:
                example = c_example

        meaning_map[w] = meaning
        example_map[w] = example

        need_m = not bool(meaning)
        need_e = not bool(example)
        if need_m or need_e:
            needs_net.append((w, need_m, need_e))

    print(f"✓ From XLSX+cache: {len(ordered) - len(needs_net)} fully covered; {len(needs_net)} need network fetch.")
    if workers < 1:
        workers = 1

    # Phase 2: network fetch missing fields concurrently
    if needs_net:
        # To avoid writing from multiple threads into sqlite, we collect results then write in main thread.
        fetched: Dict[str, Tuple[str, str]] = {}

        with ThreadPoolExecutor(max_workers=workers) as ex:
            futures = []
            for (w, need_m, need_e) in needs_net:
                futures.append(ex.submit(
                    fetch_missing_fields_for_word,
                    w, need_m, need_e,
                    retries=retries,
                    jitter_min=jitter_min,
                    jitter_max=jitter_max
                ))

            # Map futures back to words
            # Build a parallel list so we know which future is which word
            for (w, need_m, need_e), fut in zip(needs_net, futures):
                try:
                    m, e = fut.result()
                except Exception:
                    m, e = "", ""
                fetched[w] = (m or "", e or "")

        # merge + cache put
        for w, (m, e) in fetched.items():
            if not meaning_map.get(w) and m:
                meaning_map[w] = m
            if not example_map.get(w) and e:
                example_map[w] = e

            # write to cache: store the best known values (can include XLSX too)
            cache_put(cache_conn, w, meaning_map[w] or "", example_map[w] or "")

        cache_conn.commit()

    # Phase 3: final fallbacks (AI/heuristic, article sentence)
    filled_meaning_ai = 0
    filled_example_article = 0

    for i, w in enumerate(ordered, 1):
        _set_progress(status="running", current=i - 1, total=len(ordered), word=w)

        meaning = meaning_map.get(w, "") or ""
        example = example_map.get(w, "") or ""

        ai = None

        # meaning fallback: AI/heuristic if still missing
        if not meaning:
            ctx = get_surrounding_context(article_text, w, window=2, max_chars=900)
            if callable(generate_meaning_example_with_context):
                try:
                    ai = generate_meaning_example_with_context(w, ctx)
                except Exception:
                    ai = None

            if ai and (ai.get("meaning") or "").strip():
                meaning = (ai.get("meaning") or "").strip()
                filled_meaning_ai += 1
            else:
                meaning = heuristic_ai_meaning(w, ctx, example) or ""
                if meaning:
                    filled_meaning_ai += 1

        # example fallback: AI example if AI ran and has one
        if not example and ai and (ai.get("example") or "").strip():
            example = (ai.get("example") or "").strip()

        # still missing example: article sentence
        if not example:
            fallback = find_sentence_as_example(article_text, w)
            if fallback:
                example = fallback
                filled_example_article += 1

        # final meaning fallback message
        if not meaning:
            if example:
                meaning = (
                    "No standard dictionary definition found; likely a proper noun, name, "
                    "or rare term. See the example sentence for context."
                )
            else:
                meaning = (
                    "No standard dictionary definition found; likely a proper noun, name, "
                    "or rare term. See the article for context."
                )

        rows.append((w, meaning, example))
        _set_progress(status="running", current=i, total=len(ordered), word=w)

    # write CSV
    try:
        with open(out_csv, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["word", "meaning", "example"])
            writer.writerows(rows)
        print(f"\n✓ Success! Wrote {out_csv}")
        if filled_meaning_ai:
            print(f"  (AI/heuristic meaning filled: {filled_meaning_ai})")
        if filled_example_article:
            print(f"  (Article fallback examples filled: {filled_example_article})")
        _set_progress(status="done", current=len(ordered), total=len(ordered), word="")
    except PermissionError:
        ts = time.strftime("%Y%m%d_%H%M%S")
        stem, dot, ext = out_csv.partition(".")
        alt = f"{stem}_{ts}.{ext}" if dot else f"{out_csv}_{ts}"
        with open(alt, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["word", "meaning", "example"])
            writer.writerows(rows)
        print(f"WARNING: Could not write {out_csv} (file in use). Wrote {alt} instead.")
        _set_progress(status="done", current=len(ordered), total=len(ordered), word="")

    try:
        cache_conn.close()
    except Exception:
        pass


# ----------------------------
# CLI
# ----------------------------
def main():
    ap = argparse.ArgumentParser(description="Build vocabulary CSV from article + selection list (XLSX+cache+concurrency).")
    ap.add_argument("--article", default="sat.txt", help="Path to the article .txt")
    ap.add_argument("--select", default="20241226.txt", help="Path to the selection list .txt")
    ap.add_argument("--out_words", default="satword.txt", help="Output words list")
    ap.add_argument("--out_csv", default="words.csv", help="Output CSV")
    ap.add_argument("--lookup_xlsx", default="words_with_examples.xlsx", help="XLSX lookup file")
    ap.add_argument("--xlsx_sheet", default="All_Words", help="Sheet name in XLSX")

    # new speed options
    ap.add_argument("--cache_db", default="vocab_cache.sqlite", help="SQLite cache DB path")
    ap.add_argument("--workers", type=int, default=8, help="Concurrent network workers (6-12 recommended)")
    ap.add_argument("--retries", type=int, default=2, help="Retries on transient failures")
    ap.add_argument("--jitter_min", type=float, default=0.05, help="Min jitter sleep before each request")
    ap.add_argument("--jitter_max", type=float, default=0.25, help="Max jitter sleep before each request")

    args = ap.parse_args()

    build_vocab(
        args.article,
        args.select,
        args.out_words,
        args.out_csv,
        args.lookup_xlsx,
        args.xlsx_sheet,
        cache_db=args.cache_db,
        workers=args.workers,
        retries=args.retries,
        jitter_min=args.jitter_min,
        jitter_max=args.jitter_max,
    )
    print(f"Done. Wrote {args.out_words} and {args.out_csv}")


if __name__ == "__main__":
    main()
